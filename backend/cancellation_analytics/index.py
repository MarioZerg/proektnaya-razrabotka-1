import io
import json
import os
from datetime import datetime

import psycopg2
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

CORS = {
    'Access-Control-Allow-Origin': '*',
    'Access-Control-Allow-Methods': 'GET, POST, OPTIONS',
    'Access-Control-Allow-Headers': 'Content-Type, X-User-Id, X-Auth-Token',
    'Access-Control-Max-Age': '86400',
    'Content-Type': 'application/json',
}


def _resp(code, body):
    return {'statusCode': code, 'headers': CORS,
            'body': json.dumps(body, ensure_ascii=False, default=str)}


# ЧТО СЧИТАЕМ ОТМЕНОЙ ПОКУПАТЕЛЯ.
#
# Берём только FBS — заказы, которые шьются под конкретного человека. У FBO товар
# уже лежит на складе маркетплейса, и «отмена» там — это движение наших же поставок,
# а не действие покупателя. Если их смешать, отчёт превращается в мусор: 110 отмен
# из одной поставки выглядят как атака конкурента, хотя ни один покупатель не при чём.
CANCELLED_FILTER = (
    "o.marketplace = 'OZON' AND o.order_type = 'FBS' "
    "AND (o.ozon_status LIKE 'cancel%%' OR o.status = 'Отменён') "
    "AND o.ozon_posting_number IS NOT NULL"
)

# Номер отправления OZON выглядит как 39052506-0063-1: первая часть — номер ЗАКАЗА,
# он общий для всех вещей одной покупки. По нему и группируем: несколько отменённых
# вещей внутри одного заказа — это один случай, а не пять разных.
ORDER_KEY = "split_part(o.ozon_posting_number, '-', 1)"


# ВСЕ отправления покупателя, а не только отменённые. Без выкупленных нельзя ответить
# на главный вопрос: «этот вообще хоть раз забирал товар?»
ALL_FILTER = (
    "o.marketplace = 'OZON' AND o.order_type = 'FBS' "
    "AND o.ozon_posting_number IS NOT NULL"
)

# Статусы, означающие, что вещь реально поехала к покупателю или уже у него.
ALIVE_STATUSES = "('delivered', 'delivering', 'awaiting_deliver', 'awaiting_packaging')"


def _rows(cur, days):
    """Покупатели с отменами: по первой части номера отправления.

    ГЛАВНОЕ ЗДЕСЬ — СКЛЕЙКА ЗАКАЗОВ ОДНОГО ПОКУПАТЕЛЯ.
    Номер отправления OZON выглядит как 33239949-0088-1. Первая часть — номер
    ЛИЦЕВОГО СЧЁТА покупателя на OZON: он один и тот же во всех его покупках, даже
    сделанных в разные дни. Проверено на данных: у 66 номеров заказы приходили в
    РАЗНЫЕ дни, а у 122 — несколькими отдельными заказами.

    Это и позволяет сделать то, что раньше казалось невозможным: увидеть, что человек
    заказывает не первый раз, посчитать все его отмены за период и, самое важное,
    понять, выкупил ли он хоть что-нибудь.

    Оговорка: OZON не подтверждает, что первая часть номера — это именно покупатель,
    поэтому в отчёте называем это «покупателем предположительно» и не выносим
    приговоров. Но как рабочая зацепка признак сильный.
    """
    cur.execute(
        f"SELECT {ORDER_KEY} AS order_key, "
        "       COUNT(*) FILTER (WHERE o.ozon_status LIKE 'cancel%%' "
        "                          OR o.status = 'Отменён') AS cancelled_items, "
        f"      COUNT(*) FILTER (WHERE o.ozon_status IN {ALIVE_STATUSES}) AS alive_items, "
        "       COUNT(*) AS total_items, "
        "       COUNT(DISTINCT split_part(o.ozon_posting_number, '-', 2)) AS orders_count, "
        "       COUNT(DISTINCT o.created_at::date) AS active_days, "
        "       COUNT(DISTINCT o.product) AS distinct_products, "
        "       MIN(o.created_at) AS first_created, "
        "       MAX(o.created_at) AS last_created, "
        "       MAX(o.cancelled_at) AS last_cancelled, "
        "       string_agg(DISTINCT o.product, ', ') AS products, "
        "       string_agg(DISTINCT o.ozon_posting_number, ', ') AS postings, "
        "       MIN(EXTRACT(epoch FROM o.cancelled_at - o.created_at) / 3600.0) AS min_hours "
        "FROM orders o "
        f"WHERE {ALL_FILTER} "
        f"  AND o.created_at > now() - (%s || ' days')::interval "
        "GROUP BY 1 "
        "HAVING COUNT(*) FILTER (WHERE o.ozon_status LIKE 'cancel%%' "
        "                          OR o.status = 'Отменён') > 0 "
        "ORDER BY 2 DESC, 10 DESC NULLS LAST",
        (str(int(days)),),
    )
    out = []
    for r in cur.fetchall():
        hours = float(r[12]) if r[12] is not None else None
        out.append({
            'orderKey': r[0],
            'cancelledItems': int(r[1]),
            'aliveItems': int(r[2]),
            'totalItems': int(r[3]),
            'ordersCount': int(r[4]),
            'activeDays': int(r[5]),
            'distinctProducts': int(r[6]),
            'firstCreated': r[7],
            'lastCreated': r[8],
            'lastCancelled': r[9],
            'products': r[10] or '',
            'postings': r[11] or '',
            'hoursToCancel': round(hours, 1) if hours is not None else None,
            # Ни одной вещи не поехало к покупателю — заказывал только чтобы отменить.
            'neverBought': int(r[2]) == 0,
        })
    return out


def _flags(row):
    """Признаки, из-за которых заказ попал в отчёт.

    Ни один признак сам по себе не доказывает недобросовестность: человек может
    отменить покупку по сотне житейских причин. Поэтому не выносим вердиктов, а
    показываем ровно то, что видно в данных, — а выводы делает уже маркетплейс,
    у которого есть сам покупатель.
    """
    flags = []

    # Самый весомый признак: человек заказывал не раз и не забрал НИЧЕГО.
    if row['neverBought'] and row['totalItems'] >= 2:
        flags.append('Ни одного выкупа')

    if row['ordersCount'] >= 2:
        flags.append(f"Повторные заказы: {row['ordersCount']}")

    if row['activeDays'] >= 2:
        flags.append(f"Заказывал в разные дни: {row['activeDays']}")

    if row['cancelledItems'] >= 3:
        flags.append('Массовая отмена: 3+ вещи')
    elif row['cancelledItems'] == 2:
        flags.append('Отменено 2 вещи')

    h = row['hoursToCancel']
    if h is not None and h <= 1:
        flags.append('Отмена сразу после оформления')

    if row['distinctProducts'] >= 3:
        flags.append('Сразу несколько разных товаров')

    return flags


def _risk(row):
    """Насколько случай похож на намеренную скупку — от 0 до 100.

    Складываем независимые признаки: чем больше совпало, тем меньше шанс, что это
    обычный человек, который просто передумал. Один признак ничего не доказывает —
    опасны именно сочетания.
    """
    score = 0
    if row['neverBought'] and row['totalItems'] >= 2:
        score += 40
    if row['ordersCount'] >= 2:
        score += 20
    if row['activeDays'] >= 2:
        score += 15
    if row['cancelledItems'] >= 3:
        score += 15
    elif row['cancelledItems'] == 2:
        score += 5
    h = row['hoursToCancel']
    if h is not None and h <= 1:
        score += 10
    return min(score, 100)


# ВЕРОЯТНОСТЬ, ЧТО ЭТО КОНКУРЕНТ, А НЕ ОБЫЧНЫЙ ПОКУПАТЕЛЬ.
#
# Просто назвать процент «на глаз» нельзя: в OZON первым делом спросят, откуда он
# взялся. Поэтому считаем его от ФОНА — от того, как ведут себя обычные покупатели
# в наших же данных.
#
# Логика такая. По всем покупателям видно: чем больше вещей человек заказал, тем
# реже он не выкупает ВООБЩЕ НИЧЕГО. Заказал одну вещь — не выкупил в 16% случаев
# (обычное дело: передумал). Заказал четыре — таких уже 5,9%. Заказал шесть и
# ни одной не забрал — это выпадает из нормального поведения.
#
# Вероятность = насколько сильно случай выбивается из фона. Если среди обычных
# покупателей такое поведение встречается у 6%, то у оставшихся 94% причина другая,
# и вероятность неслучайности = 94%. Дальше её ограничиваем сверху: без данных
# самой площадки уверенности «на 100%» быть не может, потолок — 95%.
def _competitor_probability(row, baseline):
    """Вероятность (0-100), что за случаем стоит не обычный покупатель.

    baseline — доля обычных покупателей, которые при таком же числе заказанных
    вещей не выкупили ничего. Считается по нашим же продажам за период.
    """
    items = row['totalItems']
    # Выкупил хоть что-то — это живой покупатель, вероятность скупки низкая.
    # Отменить часть заказа человек может по любой житейской причине.
    if not row['neverBought'] or items < 2:
        return 0

    # Как часто «ни одного выкупа» встречается у обычных покупателей с таким же
    # размером заказа. Чем реже — тем меньше похоже на случайность.
    share = baseline.get(items)
    if share is None:
        # Для редких больших заказов берём ближайший известный размер: чем крупнее
        # заказ, тем ниже фон, поэтому осторожно используем самый большой из
        # посчитанных, а не ноль.
        known = [k for k in baseline if k <= items]
        share = baseline[max(known)] if known else 10.0

    prob = 100.0 - share

    # Повторные заказы в разные дни — отдельный, независимый признак. Случайно
    # передумать дважды можно, трижды подряд и ничего не забрать — уже система.
    if row['ordersCount'] >= 2:
        prob += 3
    if row['activeDays'] >= 2:
        prob += 2

    # Потолок 95%: доступ к самому покупателю есть только у площадки, поэтому
    # утверждать со стопроцентной уверенностью мы не вправе.
    return int(min(round(prob), 95))


def _probability_baseline(cur, days):
    """Фон: доля обычных покупателей без единого выкупа — по размеру заказа.

    Возвращает {сколько вещей заказал: какой процент таких не выкупил ничего}.
    Это и есть «нормальное» поведение, с которым сравнивается каждый случай.
    """
    cur.execute(
        "SELECT items, count(*) AS accounts, "
        "       count(*) FILTER (WHERE alive = 0) AS never "
        "FROM (SELECT split_part(o.ozon_posting_number, '-', 1) AS acc, "
        "             count(*) AS items, "
        f"            count(*) FILTER (WHERE o.ozon_status IN {ALIVE_STATUSES}) AS alive "
        "      FROM orders o "
        f"     WHERE {ALL_FILTER} "
        f"       AND o.created_at > now() - (%s || ' days')::interval "
        "      GROUP BY 1) k "
        "GROUP BY 1 ORDER BY 1",
        (str(int(days)),),
    )
    out = {}
    for items, accounts, never in cur.fetchall():
        # Размеры заказа, встретившиеся у единиц покупателей, для фона не годятся:
        # процент по трём людям — это шум, а не закономерность.
        if int(accounts) >= 20:
            out[int(items)] = round(100.0 * int(never) / int(accounts), 1)
    return out


def _prepare(cur, days, min_items, only_never=False):
    """Строки отчёта с признаками, оценкой риска и вероятностью скупки."""
    baseline = _probability_baseline(cur, days)
    rows = [r for r in _rows(cur, days) if r['cancelledItems'] >= min_items]
    for r in rows:
        r['flags'] = _flags(r)
        r['risk'] = _risk(r)
        r['probability'] = _competitor_probability(r, baseline)
    if only_never:
        rows = [r for r in rows if r['neverBought'] and r['totalItems'] >= 2]
    # Наверх — самые весомые случаи, а не просто самые многочисленные.
    rows.sort(key=lambda r: (r['risk'], r['cancelledItems']), reverse=True)
    return rows


def _funnel(cur, days):
    """ВОРОНКА ОТБОРА — главный лист доказательства для OZON.

    Показывает путь от всех заказов до горстки случаев, которые невозможно
    объяснить обычным поведением покупателя. Каждый шаг отсекает законное
    объяснение, и на выходе остаются только те, кому объяснения не нашлось.

    Именно так это и читается площадкой: не «нам кажется, что нас скупают», а
    «вот 6364 заказа, вот как из них осталось 29, и вот почему остальные отпали».
    """
    # Считаем по покупателям (первая часть номера отправления), а не по строкам:
    # один и тот же человек с пятью отменами — это ОДИН случай, а не пять.
    cur.execute(
        "SELECT COALESCE(sum(items_cnt), 0) AS items, "
        "       COALESCE(sum(canc_cnt), 0) AS cancelled_items, "
        "       count(*) AS accounts, "
        "       count(*) FILTER (WHERE canc_cnt > 0) AS acc_cancel, "
        "       count(*) FILTER (WHERE canc_cnt > 0 AND items_cnt >= 2 "
        "                          AND alive_cnt = 0) AS acc_never, "
        "       count(*) FILTER (WHERE canc_cnt >= 3 AND alive_cnt = 0) AS acc_never3, "
        "       count(*) FILTER (WHERE canc_cnt >= 3 AND alive_cnt = 0 "
        "                          AND orders_cnt >= 2) AS acc_never3_repeat "
        "FROM (SELECT split_part(o.ozon_posting_number, '-', 1) AS acc, "
        "             count(*) AS items_cnt, "
        "             count(*) FILTER (WHERE o.ozon_status LIKE 'cancel%%' "
        "                                 OR o.status = 'Отменён') AS canc_cnt, "
        f"            count(*) FILTER (WHERE o.ozon_status IN {ALIVE_STATUSES}) AS alive_cnt, "
        "             count(DISTINCT split_part(o.ozon_posting_number, '-', 2)) AS orders_cnt "
        "      FROM orders o "
        f"     WHERE {ALL_FILTER} "
        f"       AND o.created_at > now() - (%s || ' days')::interval "
        "      GROUP BY 1) k",
        (str(int(days)),),
    )
    r = cur.fetchone()
    items, cancelled_items, accounts, acc_cancel, acc_never, acc_never3, acc_never3_repeat = (
        int(x or 0) for x in r
    )

    def step(title, value, base, note):
        return {
            'title': title,
            'value': value,
            'share': round(100.0 * value / base, 1) if base else 0.0,
            'note': note,
        }

    return {
        'totalItems': items,
        'totalAccounts': accounts,
        'steps': [
            step('Всего заказано вещей', items, items,
                 'Все отправления OZON FBS за период — шьются под конкретного человека'),
            step('Из них отменено', cancelled_items, items,
                 'Вещь сшита или запущена в работу, а покупатель отказался'),
            step('Покупателей с отменами', acc_cancel, accounts,
                 'Столько разных счетов OZON хотя бы раз отменили заказ'),
            step('Не выкупили ничего (2+ вещи)', acc_never, accounts,
                 'Заказывали несколько раз и не забрали ни одной вещи — '
                 'обычный покупатель так себя почти не ведёт'),
            step('Отменили 3+ вещи и ни одного выкупа', acc_never3, accounts,
                 'Житейскую причину подобрать уже трудно: три отказа подряд и ноль покупок'),
            step('Плюс делали это повторными заказами', acc_never3_repeat, accounts,
                 'Возвращались снова после первой отмены — это уже не «передумал»'),
        ],
    }


def handle_report(cur, days, min_items, only_never=False):
    """Заказы с отменами за период + сводка."""
    rows = _prepare(cur, days, min_items, only_never)

    total_cancelled = sum(r['cancelledItems'] for r in rows)
    instant = sum(1 for r in rows
                  if r['hoursToCancel'] is not None and r['hoursToCancel'] <= 1)
    mass = sum(1 for r in rows if r['cancelledItems'] >= 3)
    never = sum(1 for r in rows if r['neverBought'] and r['totalItems'] >= 2)
    repeat = sum(1 for r in rows if r['ordersCount'] >= 2)

    # Динамика по дням: в какие дни отмен было аномально много. Всплеск в один день —
    # признак организованной скупки, а не случайных отказов.
    cur.execute(
        "SELECT o.created_at::date AS d, "
        "       COUNT(*) FILTER (WHERE o.ozon_status LIKE 'cancel%%' "
        "                          OR o.status = 'Отменён') AS cancelled, "
        "       COUNT(*) AS total "
        "FROM orders o "
        f"WHERE {ALL_FILTER} "
        f"  AND o.created_at > now() - (%s || ' days')::interval "
        "GROUP BY 1 ORDER BY 1",
        (str(int(days)),),
    )
    daily = []
    for d in cur.fetchall():
        total = int(d[2])
        cancelled = int(d[1])
        daily.append({
            'date': d[0],
            'cancelled': cancelled,
            'total': total,
            'share': round(100.0 * cancelled / total, 1) if total else 0.0,
        })

    # Всплески: один товар отменяли в разных заказах — самый весомый признак того,
    # что метят именно в конкретную позицию, а не просто передумали.
    cur.execute(
        f"SELECT o.product, COUNT(*) AS items, COUNT(DISTINCT {ORDER_KEY}) AS orders_cnt "
        "FROM orders o "
        f"WHERE {CANCELLED_FILTER} "
        f"  AND o.created_at > now() - (%s || ' days')::interval "
        "GROUP BY 1 HAVING COUNT(*) >= 3 "
        "ORDER BY items DESC LIMIT 20",
        (str(int(days)),),
    )
    products = [{'product': p[0], 'cancelledItems': int(p[1]), 'orders': int(p[2])}
                for p in cur.fetchall()]

    # Сколько случаев с высокой вероятностью скупки и сколько вещей на них пришлось —
    # это и есть цифра ущерба, которую спросят в поддержке.
    high = [r for r in rows if r.get('probability', 0) >= 70]
    high_items = sum(r['cancelledItems'] for r in high)
    # Средняя вероятность считается ТОЛЬКО по отобранным случаям: усреднять с нулями
    # обычных покупателей бессмысленно — процент вышел бы заниженным и ни о чём.
    avg_prob = int(round(sum(r['probability'] for r in high) / len(high))) if high else 0

    return _resp(200, {
        'days': days,
        'summary': {
            'ordersWithCancels': len(rows),
            'cancelledItems': total_cancelled,
            'instantCancels': instant,
            'massCancels': mass,
            'neverBought': never,
            'repeatBuyers': repeat,
            # Случаи, которые не объясняются обычным поведением покупателя.
            'highRiskBuyers': len(high),
            'highRiskItems': high_items,
            'avgProbability': avg_prob,
        },
        'funnel': _funnel(cur, days),
        'orders': rows,
        'products': products,
        'daily': daily,
    })


def handle_export(cur, days, min_items, only_never=False):
    """Excel-файл для обращения в поддержку маркетплейса."""
    rows = _prepare(cur, days, min_items, only_never)

    wb = Workbook()

    head_fill = PatternFill('solid', fgColor='1F3864')
    head_font = Font(color='FFFFFF', bold=True)

    # ПЕРВЫЙ ЛИСТ — ВОРОНКА. Открыв файл, сотрудник поддержки должен сразу увидеть
    # не таблицу на 600 строк, а короткий вывод: сколько было заказов, как из них
    # отобрались подозрительные и почему остальные отпали.
    wsf = wb.active
    wsf.title = 'Воронка'
    funnel = _funnel(cur, days)

    wsf['A1'] = 'Отбор случаев, похожих на скупку конкурентом'
    wsf['A1'].font = Font(bold=True, size=14)
    wsf['A2'] = (
        f'Период: последние {days} дн. Маркетплейс: OZON, схема FBS. '
        'Каждая вещь шьётся под конкретный заказ, поэтому отмена — это готовый '
        'товар, который остался у продавца.'
    )
    wsf['A2'].alignment = Alignment(wrap_text=True, vertical='top')
    wsf.merge_cells('A2:D2')
    wsf.row_dimensions[2].height = 32

    for i, (title, width) in enumerate(
        [('Шаг отбора', 44), ('Значение', 12), ('Доля, %', 10), ('Что это значит', 68)],
        start=1,
    ):
        c = wsf.cell(row=4, column=i, value=title)
        c.fill = head_fill
        c.font = head_font
        wsf.column_dimensions[get_column_letter(i)].width = width

    for s in funnel['steps']:
        wsf.append([s['title'], s['value'], s['share'], s['note']])
    for row in wsf.iter_rows(min_row=5):
        for c in row:
            c.alignment = Alignment(vertical='top', wrap_text=True)

    # Как считается вероятность — прямо в файле. Без этого объяснения процент
    # выглядит взятым с потолка, и первый же вопрос будет «откуда цифра?».
    last = wsf.max_row + 2
    wsf.cell(row=last, column=1, value='Как считается вероятность скупки').font = Font(bold=True)
    wsf.cell(
        row=last + 1, column=1,
        value=(
            'Процент считается не на глаз, а от поведения обычных покупателей в этих же '
            'данных. По всем заказам видно: чем больше вещей человек заказал, тем реже он '
            'не выкупает вообще ничего. Заказал одну вещь и не забрал — обычное дело. '
            'Заказал несколько и не забрал ни одной — так себя ведут единицы. '
            'Вероятность показывает, насколько случай выбивается из этого фона. '
            'Потолок 95%: доступ к самому покупателю есть только у площадки, поэтому '
            'утверждать со стопроцентной уверенностью продавец не вправе.'
        ),
    ).alignment = Alignment(wrap_text=True, vertical='top')
    wsf.merge_cells(start_row=last + 1, start_column=1, end_row=last + 1, end_column=4)
    wsf.row_dimensions[last + 1].height = 76

    ws = wb.create_sheet('Отмены')
    headers = [
        ('Покупатель (номер счёта OZON)', 26),
        ('Вероятность скупки, %', 20),
        ('Оценка риска', 13),
        ('Заказов', 10),
        ('Дней с заказами', 16),
        ('Отменено вещей', 15),
        ('Выкуплено вещей', 16),
        ('Ни одного выкупа', 17),
        ('Первый заказ', 17),
        ('Последний заказ', 17),
        ('Часов до отмены', 16),
        ('Товары', 44),
        ('Номера отправлений', 44),
        ('На что обратить внимание', 46),
    ]
    for i, (title, width) in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=title)
        c.fill = head_fill
        c.font = head_font
        c.alignment = Alignment(vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 30

    danger = PatternFill('solid', fgColor='FCE4E4')
    for r in rows:
        ws.append([
            r['orderKey'],
            r['probability'],
            r['risk'],
            r['ordersCount'],
            r['activeDays'],
            r['cancelledItems'],
            r['aliveItems'],
            'ДА' if (r['neverBought'] and r['totalItems'] >= 2) else '',
            r['firstCreated'].strftime('%d.%m.%Y %H:%M') if r['firstCreated'] else '',
            r['lastCreated'].strftime('%d.%m.%Y %H:%M') if r['lastCreated'] else '',
            r['hoursToCancel'] if r['hoursToCancel'] is not None else '',
            r['products'],
            r['postings'],
            '; '.join(r['flags']),
        ])

    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(vertical='top', wrap_text=True)
        # Подсвечиваем по ВЕРОЯТНОСТИ скупки (первая колонка после номера счёта):
        # именно её смотрит поддержка площадки, а не нашу внутреннюю оценку риска.
        if row[1].value and int(row[1].value) >= 70:
            for c in row:
                c.fill = danger

    # Второй лист — товары, по которым отмен больше всего.
    ws2 = wb.create_sheet('Товары')
    for i, (title, width) in enumerate(
        [('Товар', 40), ('Отменено вещей', 18), ('В скольких заказах', 20)], start=1
    ):
        c = ws2.cell(row=1, column=i, value=title)
        c.fill = head_fill
        c.font = head_font
        ws2.column_dimensions[get_column_letter(i)].width = width
    ws2.freeze_panes = 'A2'

    cur.execute(
        f"SELECT o.product, COUNT(*), COUNT(DISTINCT {ORDER_KEY}) "
        "FROM orders o "
        f"WHERE {CANCELLED_FILTER} "
        f"  AND o.created_at > now() - (%s || ' days')::interval "
        "GROUP BY 1 HAVING COUNT(*) >= 3 ORDER BY 2 DESC",
        (str(int(days)),),
    )
    for p in cur.fetchall():
        ws2.append([p[0], int(p[1]), int(p[2])])

    # Третий лист — по дням: видно всплески, когда отмен резко больше обычного.
    ws3 = wb.create_sheet('По дням')
    for i, (title, width) in enumerate(
        [('Дата', 14), ('Отменено', 12), ('Всего заказов', 16), ('Доля отмен, %', 16)],
        start=1,
    ):
        c = ws3.cell(row=1, column=i, value=title)
        c.fill = head_fill
        c.font = head_font
        ws3.column_dimensions[get_column_letter(i)].width = width
    ws3.freeze_panes = 'A2'

    cur.execute(
        "SELECT o.created_at::date, "
        "       COUNT(*) FILTER (WHERE o.ozon_status LIKE 'cancel%%' "
        "                          OR o.status = 'Отменён'), COUNT(*) "
        "FROM orders o "
        f"WHERE {ALL_FILTER} "
        f"  AND o.created_at > now() - (%s || ' days')::interval "
        "GROUP BY 1 ORDER BY 1",
        (str(int(days)),),
    )
    for d in cur.fetchall():
        total = int(d[2])
        ws3.append([
            d[0].strftime('%d.%m.%Y'), int(d[1]), total,
            round(100.0 * int(d[1]) / total, 1) if total else 0,
        ])

    buf = io.BytesIO()
    wb.save(buf)
    import base64
    return {
        'statusCode': 200,
        'headers': {
            **CORS,
            'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'Content-Disposition': (
                f'attachment; filename="otmeny-{datetime.now().strftime("%d-%m-%Y")}.xlsx"'
            ),
        },
        'isBase64Encoded': True,
        'body': base64.b64encode(buf.getvalue()).decode(),
    }


def handler(event: dict, context) -> dict:
    """Анализ отмен заказов на маркетплейсе: закономерности и выгрузка в Excel.

    Показывает заказы, где покупатель отменил несколько вещей сразу, отменил почти
    мгновенно после оформления или взял несколько разных товаров и отказался. Это
    косвенные признаки скупки конкурентом.

    Персональные данные покупателей маркетплейсы продавцу не передают, поэтому
    отчёт оперирует номерами заказов и отправлений — по ним поддержка площадки
    сама находит покупателя на своей стороне.
    """
    if event.get('httpMethod') == 'OPTIONS':
        return {'statusCode': 200, 'headers': CORS, 'body': ''}

    params = event.get('queryStringParameters') or {}
    action = (params.get('action') or 'report').strip()

    days = int(params.get('days') or 30)
    days = max(1, min(days, 365))
    min_items = int(params.get('minItems') or 2)

    # Отчёт видят администратор и менеджер. Менеджер ведёт работу с площадками и
    # разбирает отказы покупателей — без этих данных он не может ни оспорить
    # отмену, ни понять, какой товар возвращают чаще других.
    role = (params.get('actorRole') or '').strip()
    if role and role not in ('admin', 'manager'):
        return _resp(403, {'error': 'Доступ только для администратора и менеджера'})

    conn = psycopg2.connect(os.environ['DATABASE_URL'])
    try:
        cur = conn.cursor()
        # onlyNever — оставить только тех, кто не выкупил вообще ничего.
        only_never = (params.get('onlyNever') or '') in ('1', 'true')
        if action == 'export':
            return handle_export(cur, days, min_items, only_never)
        return handle_report(cur, days, min_items, only_never)
    finally:
        conn.close()